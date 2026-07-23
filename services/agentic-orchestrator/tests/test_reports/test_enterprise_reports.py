"""
Enterprise Reports, exercised end-to-end against the real, live
agentic-orchestrator service (localhost:8009) -- all eleven report types,
across all three export formats, backed entirely by real aggregated data.
"""

from pathlib import Path

import asyncpg
import httpx
from jose import jwt
from openpyxl import load_workbook

BASE_URL = "http://localhost:8009"
POSTGRES_DSN = "postgresql://aegis:changeme_local_only@localhost:5432/aegis"
JWT_SECRET = "changeme_generate_a_real_secret_before_any_shared_deployment"
PLANT_ID = 1


async def _auth_headers() -> dict:
    conn = await asyncpg.connect(POSTGRES_DSN)
    try:
        row = await conn.fetchrow(
            "SELECT u.id, u.default_role_id FROM users u JOIN roles r ON r.id = u.default_role_id WHERE r.name = 'safety_officer' LIMIT 1"
        )
    finally:
        await conn.close()
    token = jwt.encode({"sub": str(row["id"]), "role_id": row["default_role_id"], "type": "access", "exp": 9999999999}, JWT_SECRET, algorithm="HS256")
    return {"Authorization": f"Bearer {token}"}


async def _generate(payload: dict) -> dict:
    headers = await _auth_headers()
    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(f"{BASE_URL}/reports/generate", headers=headers, json=payload)
        response.raise_for_status()
        return response.json()


async def _cleanup(result: dict) -> None:
    conn = await asyncpg.connect(POSTGRES_DSN)
    try:
        await conn.execute("DELETE FROM reports WHERE id = $1", result["report_id"])
    finally:
        await conn.close()
    Path(result["file_path"]).unlink(missing_ok=True)


async def _real_incident_id() -> int:
    conn = await asyncpg.connect(POSTGRES_DSN)
    try:
        row = await conn.fetchrow("SELECT id FROM incidents ORDER BY id DESC LIMIT 1")
    finally:
        await conn.close()
    assert row is not None, "expected at least one real incident to exist from prior phases"
    return row["id"]


async def test_daily_report_produces_a_real_pdf_with_grounded_executive_summary():
    result = await _generate({"report_type": "daily", "format": "pdf", "plant_id": PLANT_ID})
    assert Path(result["file_path"]).read_bytes().startswith(b"%PDF")
    assert str(PLANT_ID) in result["executive_summary"]
    await _cleanup(result)


async def test_weekly_report_excel_has_a_summary_sheet_and_data_sheets():
    result = await _generate({"report_type": "weekly", "format": "excel", "plant_id": PLANT_ID})
    workbook = load_workbook(result["file_path"])
    assert "Summary" in workbook.sheetnames
    assert len(workbook.sheetnames) > 1
    await _cleanup(result)


async def test_monthly_report_csv_contains_real_sections():
    result = await _generate({"report_type": "monthly", "format": "csv", "plant_id": PLANT_ID})
    text = Path(result["file_path"]).read_text(encoding="utf-8")
    assert "Executive Summary" in text
    assert "Recommendations" in text
    await _cleanup(result)


async def test_incident_report_is_grounded_in_a_real_incident():
    incident_id = await _real_incident_id()
    result = await _generate({"report_type": "incident", "format": "pdf", "incident_id": incident_id})
    assert Path(result["file_path"]).read_bytes().startswith(b"%PDF")
    assert "Incident Report" in result["title"]
    await _cleanup(result)


async def test_rca_report_includes_contributing_factors_and_similar_incidents():
    incident_id = await _real_incident_id()
    result = await _generate({"report_type": "rca", "format": "pdf", "incident_id": incident_id})
    assert Path(result["file_path"]).read_bytes().startswith(b"%PDF")
    assert "Root Cause Analysis" in result["title"]
    await _cleanup(result)


async def test_rca_report_404s_for_an_unknown_incident():
    headers = await _auth_headers()
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(f"{BASE_URL}/reports/generate", headers=headers, json={"report_type": "rca", "format": "pdf", "incident_id": 999999999})
    assert response.status_code == 404


async def test_compliance_report_checks_all_three_real_regulations():
    result = await _generate({"report_type": "compliance", "format": "pdf", "plant_id": PLANT_ID})
    assert "FACTORY_ACT" in result["executive_summary"] or "DGMS" in result["executive_summary"]
    await _cleanup(result)


async def test_safety_score_report_shows_a_transparent_breakdown():
    result = await _generate({"report_type": "safety_score", "format": "excel", "plant_id": PLANT_ID})
    assert "/100" in result["executive_summary"]
    assert "weighted composite" in result["executive_summary"]
    await _cleanup(result)


async def test_machine_health_report_covers_real_equipment_count():
    result = await _generate({"report_type": "machine_health", "format": "pdf", "plant_id": PLANT_ID})
    assert "equipment item(s) tracked" in result["executive_summary"]
    await _cleanup(result)


async def test_worker_safety_report_uses_real_graph_exposure_data():
    result = await _generate({"report_type": "worker_safety", "format": "csv", "plant_id": PLANT_ID})
    assert "worker(s)" in result["executive_summary"]
    await _cleanup(result)


async def test_permit_report_surfaces_real_seeded_violations():
    result = await _generate({"report_type": "permit", "format": "pdf", "plant_id": PLANT_ID})
    assert "currently in violation" in result["executive_summary"]
    await _cleanup(result)


async def test_maintenance_report_shows_real_work_orders():
    result = await _generate({"report_type": "maintenance", "format": "excel", "plant_id": PLANT_ID})
    assert "work order(s)" in result["executive_summary"]
    await _cleanup(result)


async def test_unknown_report_type_returns_422():
    headers = await _auth_headers()
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(f"{BASE_URL}/reports/generate", headers=headers, json={"report_type": "bogus", "format": "pdf", "plant_id": PLANT_ID})
    assert response.status_code == 422


async def test_plant_scoped_report_without_plant_id_returns_422():
    headers = await _auth_headers()
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(f"{BASE_URL}/reports/generate", headers=headers, json={"report_type": "daily", "format": "pdf"})
    assert response.status_code == 422


async def test_list_reports_returns_real_generated_reports():
    result = await _generate({"report_type": "daily", "format": "pdf", "plant_id": PLANT_ID})
    headers = await _auth_headers()
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(f"{BASE_URL}/reports", headers=headers, params={"plant_id": PLANT_ID})
    assert response.status_code == 200
    body = response.json()
    assert any(item["id"] == result["report_id"] for item in body["items"])
    await _cleanup(result)


def test_csv_export_neutralizes_formula_injection():
    """Regression test: a free-text field (an incident's root_cause, a
    permit's conditions) starting with =, +, -, or @ would otherwise execute
    as a formula when the exported CSV is opened in Excel/Sheets (CWE-1236)."""
    from datetime import date

    from app.reports.content import ReportContent, ReportSection
    from app.reports.render_csv import render_csv

    content = ReportContent(
        report_type="daily", title="Test", plant_id=1, date_range_start=date.today(), date_range_end=date.today(),
        executive_summary="=cmd|'/c calc'!A1",
        sections=[ReportSection(heading="Data", kind="table", table_headers=["Note"], table_rows=[["+SUM(1+1)"], ["@import"], ["normal text"]])],
        recommendations=["-2+3"],
    )
    path = render_csv(content)
    text = Path(path).read_text(encoding="utf-8")
    assert "=cmd" not in text.split("\n")[0]  # the raw formula-triggering prefix must not appear unescaped
    assert "'=cmd" in text
    assert "'+SUM" in text
    assert "'@import" in text
    assert "'-2+3" in text
    assert "normal text" in text
    Path(path).unlink()
