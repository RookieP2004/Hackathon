"""
Renders any ReportContent into a real .xlsx workbook (openpyxl) -- a Summary
sheet (executive summary + recommendations) plus one sheet per table
section, so the same aggregated data that goes into the PDF is also
available as a genuine spreadsheet for further analysis.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.reports.content import ReportContent
from app.reports.render_pdf import REPORTS_DIR as PDF_DIR

REPORTS_DIR = PDF_DIR.parent / "enterprise"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _sheet_name(heading: str, used: set[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub("", heading)[:28] or "Sheet"
    candidate = name
    suffix = 2
    while candidate in used:
        candidate = f"{name[:25]}-{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def render_excel(content: ReportContent) -> str:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = content.title
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A2"] = f"Period: {content.date_range_start.isoformat()} to {content.date_range_end.isoformat()}"
    summary_sheet["A4"] = "Executive Summary"
    summary_sheet["A4"].font = Font(bold=True)
    summary_sheet["A5"] = content.executive_summary
    summary_sheet["A5"].alignment = summary_sheet["A5"].alignment.copy(wrap_text=True)
    summary_sheet.column_dimensions["A"].width = 100

    row = 7
    summary_sheet.cell(row=row, column=1, value="Recommendations").font = Font(bold=True)
    for rec in content.recommendations:
        row += 1
        summary_sheet.cell(row=row, column=1, value=f"- {rec}")

    used_names = {"Summary"}
    for section in content.sections:
        if section.kind != "table" or not section.table_headers:
            continue
        sheet = workbook.create_sheet(_sheet_name(section.heading, used_names))
        for col_idx, header in enumerate(section.table_headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        for row_idx, table_row in enumerate(section.table_rows or [], start=2):
            for col_idx, value in enumerate(table_row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        for col_idx in range(1, len(section.table_headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 22

    file_path = REPORTS_DIR / f"{content.report_type}-{content.plant_id or 'all'}-{int(datetime.now(timezone.utc).timestamp())}.xlsx"
    workbook.save(str(file_path))
    return str(file_path)
